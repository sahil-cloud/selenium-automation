from selenium import webdriver
from time import sleep
import openpyxl
from selenium.webdriver.common.keys import Keys

path = 'C:\\python\\jarvis\\chromedriver.exe'
excel = 'C:\\python\\jarvis\\Selenium\\Book1.xlsx'

# giving path to the driver
driver = webdriver.Chrome(path)

# opening the search results
driver.get('https://en-gb.facebook.com/login/')

driver.maximize_window()
sleep(3)

workbook = openpyxl.load_workbook(excel)
sheet = workbook.active

n_rows = sheet.max_row
print(n_rows)

for r in range(2,n_rows+1):
    user_data = sheet.cell(row=r,column=1).value
    pass_data = sheet.cell(row=r,column=2).value

    username = driver.find_element_by_id('email')
    password = driver.find_element_by_id('pass')

    username.clear()
    username.send_keys(user_data)
    password.clear()
    password.send_keys(pass_data)

    button = driver.find_element_by_id('loginbutton')
    button.click()
    sleep(5)
    # validation
    if driver.title == 'Facebook':
        print('succesful')
        sheet.cell(row=r,column=3).value = 'Passed'
        workbook.save(excel)
        break
    else:
        print('unsuceess')
        sheet.cell(row=r,column=3).value = 'failed'
        workbook.save(excel)

workbook.close()
driver.close()
